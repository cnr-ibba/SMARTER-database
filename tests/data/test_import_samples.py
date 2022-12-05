#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May  4 15:26:46 2021

@author: Paolo Cozzi <paolo.cozzi@ibba.cnr.it>
"""

import unittest
import pathlib
import tempfile

from openpyxl import Workbook
from click.testing import CliRunner
from unittest.mock import patch, PropertyMock
from pycountry.db import Data as KnownCountry

from src.data.import_samples import main as import_samples, find_country
from src.features.smarterdb import Dataset, SampleSheep, SEX
from src.features.utils import UnknownCountry

from ..common import (
    MongoMockMixin, SmarterIDMixin, SupportedChipMixin)


class TestImportSamples(
        SmarterIDMixin, SupportedChipMixin, MongoMockMixin, unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        # getting destination dataset
        cls.dst_dataset = Dataset.objects.get(file="test.zip")

        # create a src dataset
        cls.src_dataset = Dataset(
            file="test2.zip",
            country="Italy",
            species="Sheep",
            contents=[
                "metadata.xlsx"
            ],
            type_=["background", "genotypes"]
        )
        cls.src_dataset.save()

        # create a workbook
        cls.workbook = Workbook()
        cls.sheet = cls.workbook.active

        # adding header
        cls.sheet.cell(row=1, column=1, value="Code")
        cls.sheet.cell(row=1, column=2, value="Id")
        cls.sheet.cell(row=1, column=3, value="Country")
        cls.sheet.cell(row=1, column=4, value="Sex")
        cls.sheet.cell(row=1, column=5, value="Alias")
        cls.sheet.cell(row=1, column=6, value="Species")

        # adding values
        cls.sheet.cell(row=2, column=1, value="TEX_IT")
        cls.sheet.cell(row=2, column=2, value="test-1")
        cls.sheet.cell(row=2, column=3, value="Italy")
        cls.sheet.cell(row=2, column=4, value="BHO")
        cls.sheet.cell(row=2, column=5, value="test-one")
        cls.sheet.cell(row=2, column=6, value="Ovis ammon")

        # adding values
        cls.sheet.cell(row=3, column=1, value="TEX_IT")
        cls.sheet.cell(row=3, column=2, value="test-2")
        cls.sheet.cell(row=3, column=3, value="SPAIN")
        cls.sheet.cell(row=3, column=4, value="F")
        cls.sheet.cell(row=3, column=5, value="test-two")
        cls.sheet.cell(row=3, column=6, value="Ovis orientalis")

        # adding values
        cls.sheet.cell(row=4, column=1, value="TEX_IT")
        cls.sheet.cell(row=4, column=2, value="test-3")
        cls.sheet.cell(row=4, column=3, value="Italy")
        cls.sheet.cell(row=4, column=4, value="F")
        # no alias for this samples
        cls.sheet.cell(row=4, column=6, value="Ovis aries")

    def setUp(self):
        self.runner = CliRunner()

        # add a sample in database
        self.sample = SampleSheep(
            original_id="test-1",
            country="Italy",
            breed="Texel",
            breed_code="TEX",
            dataset=self.dst_dataset,
            type_="background",
            chip_name=self.chip_name,
        )
        self.sample.save()

    def tearDown(self):
        SampleSheep.objects.delete()

        super().tearDown()

    def test_help(self):
        result = self.runner.invoke(import_samples, ["--help"])
        self.assertEqual(0, result.exit_code)
        self.assertIn('Usage: main', result.output)

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_mandatory(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            result = self.runner.invoke(
                import_samples,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--code_column",
                    "Code",
                    "--country_column",
                    "Country",
                    "--id_column",
                    "Id",
                    "--chip_name",
                    self.chip_name
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            # One record is already present, two new
            self.assertEqual(SampleSheep.objects.count(), 3)

            # get the new sample
            sample = SampleSheep.objects.get(original_id="test-2")
            self.assertEqual(sample.smarter_id, "ESOA-TEX-000000002")

            # this will be the default value, since I haven't specified a
            # species column in this test
            self.assertEqual(sample.species, "Ovis aries")

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_with_sex(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            result = self.runner.invoke(
                import_samples,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--code_column",
                    "Code",
                    "--country_column",
                    "Country",
                    "--id_column",
                    "Id",
                    "--sex_column",
                    "Sex",
                    "--chip_name",
                    self.chip_name
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            # One record is already present, two new
            self.assertEqual(SampleSheep.objects.count(), 3)

            # first sampla wasn't been updated. Unknown sex
            self.sample.reload()
            self.assertIsNone(self.sample.sex)

            # get the new sample
            sample = SampleSheep.objects.get(original_id="test-2")
            self.assertEqual(sample.smarter_id, "ESOA-TEX-000000002")
            self.assertEqual(sample.sex, SEX.FEMALE)
            # default value for this sample
            self.assertEqual(sample.species, "Ovis aries")

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_same_country_and_code(self, my_working_dir):
        """Apply the same country and code to all samples"""

        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            result = self.runner.invoke(
                import_samples,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--code_all",
                    "TEX",
                    "--country_all",
                    "Italy",
                    "--id_column",
                    "Id",
                    "--chip_name",
                    self.chip_name
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            # One record is already present, two new
            self.assertEqual(SampleSheep.objects.count(), 3)

            # get the new sample: country will be always Italy
            sample = SampleSheep.objects.get(original_id="test-2")
            self.assertEqual(sample.smarter_id, "ITOA-TEX-000000002")
            # default value with no species
            self.assertEqual(sample.species, "Ovis aries")

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_alias(self, my_working_dir):
        """Test importing samples with aliases"""

        # alias need to be defined in original samples, to get the same
        # entity
        self.sample.alias = "test-one"
        self.sample.save()

        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            result = self.runner.invoke(
                import_samples,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--code_column",
                    "Code",
                    "--country_column",
                    "Country",
                    "--id_column",
                    "Id",
                    "--alias_column",
                    "Alias",
                    "--chip_name",
                    self.chip_name
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            # One record is already present, two new
            self.assertEqual(SampleSheep.objects.count(), 3)

            # get the old sample
            sample = SampleSheep.objects.get(original_id="test-1")
            self.assertEqual(sample.smarter_id, "ITOA-TEX-000000001")
            self.assertEqual(sample.alias, "test-one")

            # get the new sample
            sample = SampleSheep.objects.get(original_id="test-2")
            self.assertEqual(sample.smarter_id, "ESOA-TEX-000000002")
            self.assertEqual(sample.alias, "test-two")
            self.assertEqual(sample.species, "Ovis aries")

            # get the sample with no alias
            sample = SampleSheep.objects.get(original_id="test-3")
            self.assertEqual(sample.smarter_id, "ITOA-TEX-000000003")
            self.assertIsNone(sample.alias)
            self.assertEqual(sample.species, "Ovis aries")

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_alias_skip(self, my_working_dir):
        """Test importing samples skipping by aliases"""

        # alias need to be defined in original samples, to get the same
        # entity
        self.sample.alias = "test-one"
        self.sample.save()

        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            result = self.runner.invoke(
                import_samples,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--code_column",
                    "Code",
                    "--country_column",
                    "Country",
                    "--id_column",
                    "Id",
                    "--alias_column",
                    "Alias",
                    "--skip_missing_alias",
                    "--chip_name",
                    self.chip_name
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            # I should have two record for samples, One already present one new
            self.assertEqual(SampleSheep.objects.count(), 2)

            # get the old sample
            sample = SampleSheep.objects.get(original_id="test-1")
            self.assertEqual(sample.smarter_id, "ITOA-TEX-000000001")
            self.assertEqual(sample.alias, "test-one")

            # get the new sample
            sample = SampleSheep.objects.get(original_id="test-2")
            self.assertEqual(sample.smarter_id, "ESOA-TEX-000000002")
            self.assertEqual(sample.alias, "test-two")
            self.assertEqual(sample.species, "Ovis aries")

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_with_species(self, my_working_dir):
        """Test importing samples with species"""

        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            result = self.runner.invoke(
                import_samples,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--code_column",
                    "Code",
                    "--country_column",
                    "Country",
                    "--id_column",
                    "Id",
                    "--chip_name",
                    self.chip_name,
                    "--species_column",
                    "Species"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            # One record is already present, two new
            self.assertEqual(SampleSheep.objects.count(), 3)

            # get the old sample. Species hasn't changed
            sample = SampleSheep.objects.get(original_id="test-1")
            self.assertEqual(sample.smarter_id, "ITOA-TEX-000000001")
            self.assertEqual(sample.species, "Ovis aries")

            # get the new sample. Species is different from default
            sample = SampleSheep.objects.get(original_id="test-2")
            self.assertEqual(sample.smarter_id, "ESOA-TEX-000000002")
            self.assertEqual(sample.species, "Ovis orientalis")


class TestFindCountry(unittest.TestCase):
    def test_get_unknown_country(self):
        test = find_country("Unknown")
        self.assertIsInstance(test, UnknownCountry)
        self.assertEqual(test.name, "Unknown")
        self.assertEqual(test.alpha_2, "UN")
        self.assertEqual(test.alpha_3, "UNK")

    def test_get_known_country(self):
        test = find_country("Italy")
        self.assertIsInstance(test, KnownCountry)
        self.assertEqual(test.name, "Italy")
        self.assertEqual(test.alpha_2, "IT")
        self.assertEqual(test.alpha_3, "ITA")


if __name__ == '__main__':
    unittest.main()
