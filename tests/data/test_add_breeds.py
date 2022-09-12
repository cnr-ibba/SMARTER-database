#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Apr 20 14:56:16 2021

@author: Paolo Cozzi <paolo.cozzi@ibba.cnr.it>
"""

import unittest
import pathlib
import tempfile

from openpyxl import Workbook
from click.testing import CliRunner
from unittest.mock import patch, PropertyMock

from src.features.smarterdb import Breed, BreedAlias, Dataset
from src.data.add_breed import main as add_breed
from src.data.import_breeds import main as import_breeds

from ..common import MongoMockMixin


class BreedMixin():
    main_function = None

    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        # need a dataset for certain tests
        cls.dataset = Dataset(
            file="test.zip",
            country="Italy",
            species="Sheep",
            contents=[
                "plinktest.map",
                "plinktest.ped",
                "breed.xlsx"
            ]
        )
        cls.dataset.save()

    @classmethod
    def tearDownClass(cls):
        Dataset.objects.delete()

        super().tearDownClass()

    def setUp(self):
        super().setUp()

        self.runner = CliRunner()

    def tearDown(self):
        # drop all created breed
        Breed.objects.delete()

        super().tearDown()

    def test_help(self):
        result = self.runner.invoke(self.main_function, ["--help"])
        self.assertEqual(0, result.exit_code)
        self.assertIn('Usage: main', result.output)


class AddBreedTest(BreedMixin, MongoMockMixin, unittest.TestCase):
    main_function = add_breed

    def test_add_breed(self):
        result = self.runner.invoke(
            self.main_function,
            [
                "--species_class",
                "sheep",
                "--name",
                "Texel",
                "--code",
                "TEX",
                "--dataset",
                "test.zip",
                "--alias",
                "TEXEL_IT",
                "--alias",
                "0"
            ]
        )

        self.assertEqual(0, result.exit_code, msg=result.exc_info)

        qs = Breed.objects()
        self.assertEqual(qs.count(), 1)

        breed = qs.get()
        aliases = [
            BreedAlias(fid=fid, dataset=self.dataset)
            for fid in ["TEXEL_IT", "0"]]
        self.assertEqual(breed.aliases, aliases)


class ImportBreedsTest(BreedMixin, MongoMockMixin, unittest.TestCase):
    main_function = import_breeds

    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        # create a workbook
        cls.workbook = Workbook()
        cls.sheet = cls.workbook.active

        # adding header
        cls.sheet.cell(row=1, column=1, value="Code")
        cls.sheet.cell(row=1, column=2, value="Name")
        cls.sheet.cell(row=1, column=3, value="Fid")
        cls.sheet.cell(row=1, column=4, value="Country")

        # adding values
        cls.sheet.cell(row=2, column=1, value="TEX")
        cls.sheet.cell(row=2, column=2, value="Texel")
        cls.sheet.cell(row=2, column=3, value="XET")
        cls.sheet.cell(row=2, column=4, value="Italy")

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_breeds(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/breed.xlsx")

            result = self.runner.invoke(
                self.main_function,
                [
                    "--species_class",
                    "sheep",
                    "--src_dataset",
                    "test.zip",
                    "--datafile",
                    "breed.xlsx",
                    "--code_column",
                    "Code",
                    "--breed_column",
                    "Name"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            qs = Breed.objects()
            self.assertEqual(qs.count(), 1)

            breed = qs.get()
            alias = BreedAlias(fid="TEX", dataset=self.dataset, country=None)
            self.assertEqual(breed.aliases, [alias])

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_breeds_force_fid(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/breed.xlsx")

            result = self.runner.invoke(
                self.main_function,
                [
                    "--species_class",
                    "sheep",
                    "--src_dataset",
                    "test.zip",
                    "--datafile",
                    "breed.xlsx",
                    "--code_column",
                    "Code",
                    "--breed_column",
                    "Name",
                    "--fid_column",
                    "Fid"

                ]
            )

            self.assertEqual(0, result.exit_code)

            qs = Breed.objects()
            self.assertEqual(qs.count(), 1)

            breed = qs.get()
            alias = BreedAlias(fid="XET", dataset=self.dataset, country=None)
            self.assertEqual(breed.aliases, [alias])

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_breeds_force_country(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/breed.xlsx")

            result = self.runner.invoke(
                self.main_function,
                [
                    "--species_class",
                    "sheep",
                    "--src_dataset",
                    "test.zip",
                    "--datafile",
                    "breed.xlsx",
                    "--code_column",
                    "Code",
                    "--breed_column",
                    "Name",
                    "--country_column",
                    "Country",
                ]
            )

            self.assertEqual(0, result.exit_code)

            qs = Breed.objects()
            self.assertEqual(qs.count(), 1)

            breed = qs.get()
            alias = BreedAlias(
                fid="TEX", dataset=self.dataset, country="Italy")
            self.assertEqual(breed.aliases, [alias])


if __name__ == '__main__':
    unittest.main()
