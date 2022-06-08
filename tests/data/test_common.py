#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Apr 30 15:45:36 2021

@author: Paolo Cozzi <paolo.cozzi@ibba.cnr.it>
"""

import pathlib
import datetime
import unittest
import tempfile

import pandas as pd
from unittest.mock import patch, PropertyMock
from openpyxl import Workbook

from src.data.common import (
    fetch_and_check_dataset, get_variant_species, get_sample_species,
    pandas_open, update_chip_name, update_sequence, update_affymetrix_record,
    update_location, update_variant, update_rs_id, update_probesets)
from src.features.smarterdb import (
    Dataset, VariantGoat, VariantSheep, SampleSheep, SampleGoat, Location,
    Probeset)

from ..common import MongoMockMixin, SmarterIDMixin, VariantsMixin

FEATURE_DATA_DIR = pathlib.Path(__file__).parents[1] / "features/data"
SCRIPTS_DATA_DIR = pathlib.Path(__file__).parents[1] / "data/data"


class CommonScriptTest(SmarterIDMixin, MongoMockMixin, unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.snpfile = FEATURE_DATA_DIR / "snplist.txt"
        cls.report = FEATURE_DATA_DIR / "finalreport.txt"
        cls.archive = "test.zip"
        cls.contents = ["snplist.txt", "finalreport.txt"]

        super().setUpClass()

    def link_files(self, working_dir):
        snpfile = working_dir / "snplist.txt"
        report = working_dir / "finalreport.txt"

        snpfile.symlink_to(self.snpfile)
        report.symlink_to(self.report)

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_get_dataset_with_stuff(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)

            # assign return value to mocked property
            my_working_dir.return_value = working_dir

            # copy test data files
            self.link_files(working_dir)

            # test stuff
            dataset, [snplist, finalreport] = fetch_and_check_dataset(
                archive=self.archive,
                contents=self.contents,
            )

            self.assertIsInstance(dataset, Dataset)
            self.assertIsInstance(snplist, pathlib.Path)
            self.assertIsInstance(finalreport, pathlib.Path)

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_file_not_in_dataset(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)

            # assign return value to mocked property
            my_working_dir.return_value = working_dir

            # copy test data files
            self.link_files(working_dir)

            # test stuff
            self.assertRaisesRegex(
                FileNotFoundError,
                r"Couldn't find '\['not_present.txt'\]'",
                fetch_and_check_dataset,
                self.archive,
                ["not_present.txt"]
            )

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_file_not_in_workdir(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)

            # assign return value to mocked property
            my_working_dir.return_value = working_dir

            # don't create test file symlinks in directory

            self.assertRaisesRegex(
                FileNotFoundError,
                r"Couldn't find '\['snplist.txt', 'finalreport.txt'\]'",
                fetch_and_check_dataset,
                self.archive,
                self.contents
            )

    def test_workingdir_not_exists(self):
        # don't create a wroking dir
        self.assertRaisesRegex(
            FileNotFoundError,
            "Could find dataset directory",
            fetch_and_check_dataset,
            self.archive,
            self.contents
        )


class GetVariantSpeciesTest(unittest.TestCase):
    def test_get_variant_sheep(self):
        VariantSpecies = get_variant_species(species="Sheep")
        self.assertEqual(VariantSpecies, VariantSheep)

    def test_get_variant_goat(self):
        VariantSpecies = get_variant_species(species="Goat")
        self.assertEqual(VariantSpecies, VariantGoat)

    def test_unmanaged_species(self):
        self.assertRaisesRegex(
            NotImplementedError,
            "not yet implemented",
            get_variant_species,
            "unmanaged"
        )


class GetSampleSpeciesTest(unittest.TestCase):
    def test_get_sample_sheep(self):
        SampleSpecies = get_sample_species(species="Sheep")
        self.assertEqual(SampleSpecies, SampleSheep)

    def test_get_sample_goat(self):
        SampleSpecies = get_sample_species(species="Goat")
        self.assertEqual(SampleSpecies, SampleGoat)

    def test_unmanaged_species(self):
        self.assertRaisesRegex(
            NotImplementedError,
            "not yet implemented",
            get_sample_species,
            "unmanaged"
        )


class PandasOpenTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        # create a workbook
        cls.workbook = Workbook()
        cls.sheet = cls.workbook.active

        # adding header
        cls.sheet.cell(row=1, column=1, value="Code")
        cls.sheet.cell(row=1, column=2, value="Name")
        cls.sheet.cell(row=1, column=3, value="Fid")
        cls.sheet.cell(row=1, column=4, value="Country")

        # adding values
        cls.sheet.cell(row=2, column=1, value="TEX")
        cls.sheet.cell(row=2, column=2, value="Texel")
        cls.sheet.cell(row=2, column=3, value="XET")
        cls.sheet.cell(row=2, column=4, value="Italy")

    def test_open_csv(self):
        data = pandas_open(SCRIPTS_DATA_DIR / "test_manifest.csv")
        self.assertIsInstance(data, pd.DataFrame)

    def test_open_xlsx(self):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            datapath = working_dir / "breed.xlsx"

            # save worksheet in temporary folder
            self.workbook.save(f"{datapath}")

            data = pandas_open(datapath)
            self.assertIsInstance(data, pd.DataFrame)

    def test_open_xls(self):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            datapath = working_dir / "breed.xls"

            # save worksheet in temporary folder
            self.workbook.save(f"{datapath}")

            data = pandas_open(datapath)
            self.assertIsInstance(data, pd.DataFrame)


class VariantUpdateTests(VariantsMixin, MongoMockMixin, unittest.TestCase):
    def setUp(self):
        self.record = VariantSheep.objects.get(
            name="250506CS3900065000002_1238.1")

        # make a copy, those are all refrences
        variant_data = self.data[0].copy()
        location_data = variant_data.pop('locations')[0]

        self.variant = VariantSheep(**variant_data)
        self.location = Location(**location_data)

    def tearDown(self):
        # reset record to its original state
        self.record.update(**self.data[0])

    def test_update_chip_name(self):
        record, updated = update_chip_name(self.variant, self.record)
        self.assertFalse(updated)

        # set a new chip_name
        self.variant.chip_name = ["test"]
        record, updated = update_chip_name(self.variant, self.record)
        self.assertTrue(updated)

    def test_update_sequence(self):
        record, updated = update_sequence(self.variant, self.record)
        self.assertFalse(updated)

        # delete sequence from variant, no update
        self.variant.sequence = None
        record, updated = update_sequence(self.variant, self.record)
        self.assertFalse(updated)

        # update sequence
        self.variant.sequence = {"test": "AGCT"}
        record, updated = update_sequence(self.variant, self.record)
        self.assertTrue(updated)

    def test_update_affymetrix_record(self):
        record, updated = update_affymetrix_record(self.variant, self.record)
        self.assertFalse(updated)

        # add affymetrix attribute
        self.variant.probeset_id = "test"
        self.variant.affy_snp_id = "test"

        record, updated = update_affymetrix_record(self.variant, self.record)
        self.assertTrue(updated)

    def test_update_location(self):
        record, updated = update_location(self.location, self.record)
        self.assertFalse(updated)

        # change location
        self.location.version = "test"
        record, updated = update_location(self.location, self.record)
        self.assertTrue(updated)

    def test_update_location_no_update(self):
        """Test a location mismatch"""
        self.location.chrom = "test"
        record, updated = update_location(self.location, self.record)
        self.assertFalse(updated)

        # no location added if the new location is older
        self.assertEqual(len(record.locations), 3)

    def test_update_location_new_location(self):
        """Test a location mismatch with a new location"""

        # now update location date
        self.location.chrom = "test"
        self.location.date = datetime.datetime(2022, 5, 27)
        record, updated = update_location(self.location, self.record)
        self.assertTrue(updated)

        # now location is updated
        self.assertEqual(len(record.locations), 3)

        location = record.get_location(
            version="Oar_v3.1", imported_from="manifest")
        self.assertEqual(location.chrom, "test")
        self.assertEqual(location.date, datetime.datetime(2022, 5, 27))

    def test_update_rs_id(self):
        record, updated = update_rs_id(self.variant, self.record)
        self.assertFalse(updated)

        # unset rs_id
        self.variant.rs_id = None
        record, updated = update_rs_id(self.variant, self.record)
        self.assertFalse(updated)

        # udpate rs_id
        self.variant.rs_id = "test"
        record, updated = update_rs_id(self.variant, self.record)
        self.assertTrue(updated)

    def test_update_probeset(self):
        record_attr = [Probeset(chip_name="test", probeset_id=["test"])]
        variant_attr = [Probeset(chip_name="test2", probeset_id=['test'])]

        # test for the same probe in a different chip
        updated = update_probesets(variant_attr, record_attr)
        self.assertTrue(updated)
        self.assertEqual(len(record_attr), 2)
        self.assertEqual(len(record_attr[0].probeset_id), 1)
        self.assertEqual(len(record_attr[1].probeset_id), 1)

        record_attr = [Probeset(chip_name="test", probeset_id=["test"])]
        variant_attr = [Probeset(chip_name="test", probeset_id=['test2'])]

        # test for a new probeset in the same chip
        updated = update_probesets(variant_attr, record_attr)
        self.assertTrue(updated)
        self.assertEqual(len(record_attr), 1)
        self.assertEqual(len(record_attr[0].probeset_id), 2)

    def test_update_variant_snp_mismatch(self):
        """No update if location.illumina_top is different from
        variant.illumina_top"""

        qs = VariantSheep.objects.filter(
            name="250506CS3900065000002_1238.1")

        self.location.illumina_top = "A/C"

        updated = update_variant(qs, self.variant, self.location)
        self.assertFalse(updated)

    def test_update_variant_chip_name(self):
        qs = VariantSheep.objects.filter(
            name="250506CS3900065000002_1238.1")
        updated = update_variant(qs, self.variant, self.location)
        self.assertFalse(updated)

        # set a new chip_name
        self.variant.chip_name = ["test"]
        updated = update_variant(qs, self.variant, self.location)
        self.assertTrue(updated)

    def test_update_variant_sequence(self):
        qs = VariantSheep.objects.filter(
            name="250506CS3900065000002_1238.1")
        updated = update_variant(qs, self.variant, self.location)
        self.assertFalse(updated)

        # update sequence
        self.variant.sequence = {"test": "AGCT"}
        updated = update_variant(qs, self.variant, self.location)
        self.assertTrue(updated)

    def test_update_variant_affymetrix_record(self):
        qs = VariantSheep.objects.filter(
            name="250506CS3900065000002_1238.1")
        updated = update_variant(qs, self.variant, self.location)
        self.assertFalse(updated)

        # add affymetrix attribute
        self.variant.probesets = [
            Probeset(chip_name="AffymetrixAxiomOviCan", probeset_id=["test"])]
        self.variant.affy_snp_id = "test"
        updated = update_variant(qs, self.variant, self.location)
        self.assertTrue(updated)

    def test_update_variant_with_location(self):
        qs = VariantSheep.objects.filter(
            name="250506CS3900065000002_1238.1")
        updated = update_variant(qs, self.variant, self.location)
        self.assertFalse(updated)

        # change location
        self.location.version = "test"
        updated = update_variant(qs, self.variant, self.location)
        self.assertTrue(updated)

    def test_update_variant_with_rs_id(self):
        qs = VariantSheep.objects.filter(
            name="250506CS3900065000002_1238.1")
        updated = update_variant(qs, self.variant, self.location)
        self.assertFalse(updated)

        # update rs_id
        self.variant.rs_id = "test"
        updated = update_variant(qs, self.variant, self.location)
        self.assertTrue(updated)


if __name__ == '__main__':
    unittest.main()
