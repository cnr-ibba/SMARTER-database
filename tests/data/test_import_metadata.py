#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Apr 30 16:19:29 2021

@author: Paolo Cozzi <paolo.cozzi@ibba.cnr.it>
"""

import unittest
import pathlib
import tempfile

from openpyxl import Workbook
from click.testing import CliRunner
from unittest.mock import patch, PropertyMock

from src.data.import_metadata import main as import_metadata
from src.features.smarterdb import Dataset, SampleSheep

from ..common import MongoMockMixin, SmarterIDMixin, SupportedChipMixin


class MetaDataMixin(SmarterIDMixin, SupportedChipMixin, MongoMockMixin):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        # getting destination dataset
        cls.dst_dataset = Dataset.objects.get(file="test.zip")

        # create a src dataset
        cls.src_dataset = Dataset(
            file="test2.zip",
            country="Italy",
            species="Sheep",
            contents=[
                "metadata.xlsx"
            ]
        )
        cls.src_dataset.save()

        # create a workbook
        cls.workbook = Workbook()
        cls.sheet = cls.workbook.active

        # adding header
        cls.sheet.cell(row=1, column=1, value="Code")
        cls.sheet.cell(row=1, column=2, value="Name")
        cls.sheet.cell(row=1, column=3, value="Country")
        cls.sheet.cell(row=1, column=4, value="Lat")
        cls.sheet.cell(row=1, column=5, value="Lon")
        cls.sheet.cell(row=1, column=6, value="Col1")
        cls.sheet.cell(row=1, column=7, value="Col 2")
        cls.sheet.cell(row=1, column=8, value="Id")
        cls.sheet.cell(row=1, column=9, value="Alias")

        # adding values
        cls.sheet.cell(row=2, column=1, value="TEX")
        cls.sheet.cell(row=2, column=2, value="Texel")
        cls.sheet.cell(row=2, column=3, value="Italy")
        cls.sheet.cell(row=2, column=4, value=45.46427)
        cls.sheet.cell(row=2, column=5, value=9.18951)
        cls.sheet.cell(row=2, column=6, value="Val1")
        cls.sheet.cell(row=2, column=7, value="Val2")
        cls.sheet.cell(row=2, column=8, value="test-1")
        cls.sheet.cell(row=2, column=9, value="test-one")

        cls.sheet.cell(row=3, column=1, value="MER")
        cls.sheet.cell(row=3, column=2, value="Merino")
        cls.sheet.cell(row=3, column=3, value="Italy")
        cls.sheet.cell(row=3, column=4, value="45.46427/46.46427")
        cls.sheet.cell(row=3, column=5, value="9.18951/10.18951")
        cls.sheet.cell(row=3, column=6, value="Val3")
        cls.sheet.cell(row=3, column=7, value="Val4")
        cls.sheet.cell(row=3, column=8, value="test-2")
        cls.sheet.cell(row=3, column=9, value="test-two")

    @classmethod
    def tearDownClass(cls):
        Dataset.objects.delete()
        SampleSheep.objects.delete()

        super().tearDownClass()

    def setUp(self):
        self.runner = CliRunner()

        # need also a sample
        self.sample1 = SampleSheep(
            original_id="test-1",
            smarter_id="ITOA-TEX-000000001",
            country="Italy",
            breed="Texel",
            breed_code="TEX",
            dataset=self.dst_dataset,
            type_="background",
            chip_name=self.chip_name,
            alias="test-one"
        )
        self.sample1.save()

        self.sample2 = SampleSheep(
            original_id="test-2",
            smarter_id="ITOA-MER-000000002",
            country="Italy",
            breed="Merino",
            breed_code="MER",
            dataset=self.dst_dataset,
            type_="background",
            chip_name=self.chip_name,
            alias="test-two"
        )
        self.sample2.save()

    def tearDown(self):
        SampleSheep.objects.delete()

        super().tearDown()

    def check_sample1_locations(self):
        self.sample1.reload()
        self.assertEqual(
            self.sample1.locations,
            {
                'type': 'MultiPoint',
                'coordinates': [[9.18951, 45.46427]]
            }
        )

    def check_sample2_locations(self):
        self.sample2.reload()
        self.assertEqual(
            self.sample2.locations,
            {
                'type': 'MultiPoint',
                'coordinates': [[9.18951, 45.46427], [10.18951, 46.46427]]
            }
        )

    def check_sample1_metadata(self):
        self.sample1.reload()
        self.assertEqual(
            self.sample1.metadata,
            {
                'col1': 'Val1',
                'col_2': 'Val2'
            }
        )

    def check_sample2_metadata(self):
        self.sample2.reload()
        self.assertEqual(
            self.sample2.metadata,
            {
                'col1': 'Val3',
                'col_2': 'Val4'
            }
        )


class TestImportMetadataCLI(MetaDataMixin, unittest.TestCase):
    def test_help(self):
        result = self.runner.invoke(import_metadata, ["--help"])
        self.assertEqual(0, result.exit_code)
        self.assertIn('Usage: main', result.output)

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_mutually_exclusive_options(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            result = self.runner.invoke(
                import_metadata,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--breed_column",
                    "Name",
                    "--id_column",
                    "Id"
                ]
            )

            self.assertEqual(2, result.exit_code)


class TestImportMetadataByBreeds(MetaDataMixin, unittest.TestCase):
    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_with_position(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # test two records in database
            self.assertEqual(SampleSheep.objects.count(), 2)

            result = self.runner.invoke(
                import_metadata,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--breed_column",
                    "Name",
                    "--latitude_column",
                    "Lat",
                    "--longitude_column",
                    "Lon"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            # check locations
            self.check_sample1_locations()
            self.check_sample2_locations()

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_with_metadata(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # test two records in database
            self.assertEqual(SampleSheep.objects.count(), 2)

            result = self.runner.invoke(
                import_metadata,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--breed_column",
                    "Name",
                    "--metadata_column",
                    "Col1",
                    "--metadata_column",
                    "Col 2"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            # check metadata
            self.check_sample1_metadata()
            self.check_sample2_metadata()


class TestImportMetadataBySamples(MetaDataMixin, unittest.TestCase):
    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_with_position(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # test two records in database
            self.assertEqual(SampleSheep.objects.count(), 2)

            result = self.runner.invoke(
                import_metadata,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--id_column",
                    "Id",
                    "--latitude_column",
                    "Lat",
                    "--longitude_column",
                    "Lon"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            # check locationss
            self.check_sample1_locations()
            self.check_sample2_locations()

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_with_metadata(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # test two records in database
            self.assertEqual(SampleSheep.objects.count(), 2)

            result = self.runner.invoke(
                import_metadata,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--id_column",
                    "Id",
                    "--metadata_column",
                    "Col1",
                    "--metadata_column",
                    "Col 2"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            # check metadata
            self.check_sample1_metadata()
            self.check_sample2_metadata()


class TestImportMetadataBySamplesNA(MetaDataMixin, unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        # setting NA values for lat and long
        cls.sheet.cell(row=1, column=4, value="Lat")
        cls.sheet.cell(row=1, column=5, value="Lon")

        cls.sheet.cell(row=2, column=4, value="NA")
        cls.sheet.cell(row=2, column=5, value="NA")

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_with_position_na(self, my_working_dir):
        """Passing NA values for coordinates doesn't raise issues"""

        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # test two records in database
            self.assertEqual(SampleSheep.objects.count(), 2)

            result = self.runner.invoke(
                import_metadata,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--id_column",
                    "Id",
                    "--latitude_column",
                    "Lat",
                    "--longitude_column",
                    "Lon",
                    "--na_values",
                    "NA"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            self.sample1.reload()
            self.assertIsNone(self.sample1.locations)

            self.check_sample2_locations()


class TestImportMetadataByAlias(MetaDataMixin, unittest.TestCase):
    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_with_position(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # test two records in database
            self.assertEqual(SampleSheep.objects.count(), 2)

            result = self.runner.invoke(
                import_metadata,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--alias_column",
                    "Alias",
                    "--latitude_column",
                    "Lat",
                    "--longitude_column",
                    "Lon"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            # check locationss
            self.check_sample1_locations()
            self.check_sample2_locations()

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_with_metadata(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/metadata.xlsx")

            # test two records in database
            self.assertEqual(SampleSheep.objects.count(), 2)

            result = self.runner.invoke(
                import_metadata,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "metadata.xlsx",
                    "--alias_column",
                    "Alias",
                    "--metadata_column",
                    "Col1",
                    "--metadata_column",
                    "Col 2"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)

            # check metadata
            self.check_sample1_metadata()
            self.check_sample2_metadata()


if __name__ == '__main__':
    unittest.main()
