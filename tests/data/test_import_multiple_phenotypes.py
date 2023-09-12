#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Sep 11 12:20:08 2023

@author: Paolo Cozzi <paolo.cozzi@ibba.cnr.it>
"""

import unittest
import pathlib
import tempfile

from openpyxl import Workbook
from click.testing import CliRunner
from unittest.mock import patch, PropertyMock

from src.data.import_multiple_phenotypes import (
    main as import_multiple_phenotypes)
from src.features.smarterdb import Dataset, SampleSheep, Phenotype

from ..common import MongoMockMixin, SmarterIDMixin, SupportedChipMixin


class PhenotypeMixin(SmarterIDMixin, SupportedChipMixin, MongoMockMixin):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        # getting destination dataset
        cls.dst_dataset = Dataset.objects.get(file="test.zip")

        # create a src dataset for genotypes
        cls.src_dataset = Dataset(
            file="test2.zip",
            country="Italy",
            species="Sheep",
            contents=[
                "phenotypes.xlsx"
            ]
        )
        cls.src_dataset.save()

        # create a workbook
        cls.workbook = Workbook()
        cls.sheet = cls.workbook.active

        # adding header
        cls.sheet.cell(row=1, column=1, value="Id")
        cls.sheet.cell(row=1, column=2, value="daily_activity_min")
        cls.sheet.cell(row=1, column=3, value="daily_distance_km")

        # adding values
        cls.sheet.cell(row=2, column=1, value="test-1")
        cls.sheet.cell(row=2, column=2, value=123)
        cls.sheet.cell(row=2, column=3, value=11.5)
        cls.sheet.cell(row=3, column=1, value="test-1")
        cls.sheet.cell(row=3, column=2, value=456)
        cls.sheet.cell(row=3, column=3, value=23.0)
        cls.sheet.cell(row=4, column=1, value="test-2")
        cls.sheet.cell(row=4, column=2, value=123)
        cls.sheet.cell(row=4, column=3, value=11.5)

    @classmethod
    def tearDownClass(cls):
        Dataset.objects.delete()
        SampleSheep.objects.delete()

        super().tearDownClass()

    def setUp(self):
        self.runner = CliRunner()

        # need also a sample
        self.sample = SampleSheep(
            original_id="test-1",
            smarter_id="ITOA-TEX-000000001",
            country="Italy",
            breed="Texel",
            breed_code="TEX",
            dataset=self.dst_dataset,
            type_="background",
            chip_name=self.chip_name,
            alias="alias-1",
        )
        self.sample.save()

    def tearDown(self):
        SampleSheep.objects.delete()

        super().tearDown()


class TestImportPhenotypeCLI(PhenotypeMixin, unittest.TestCase):
    def test_help(self):
        result = self.runner.invoke(import_multiple_phenotypes, ["--help"])
        self.assertEqual(0, result.exit_code)
        self.assertIn('Usage: main', result.output)


class TestImportPhenotypeBySamples(PhenotypeMixin, unittest.TestCase):
    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_multiple_phenotype(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/phenotypes.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            result = self.runner.invoke(
                import_multiple_phenotypes,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "phenotypes.xlsx",
                    "--id_column",
                    "Id",
                    "--column",
                    "daily_activity_min",
                    "--column",
                    "daily_distance_km",
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)
            self.sample.reload()
            self.assertIsInstance(self.sample.phenotype, Phenotype)

            reference = Phenotype(
                daily_activity_min=[123, 456],
                daily_distance_km=[11.5, 23.0]
            )

            self.assertEqual(reference, self.sample.phenotype)
