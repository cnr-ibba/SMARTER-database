#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 11 12:12:12 2021

@author: Paolo Cozzi <paolo.cozzi@ibba.cnr.it>
"""

import unittest
import pathlib
import tempfile

from openpyxl import Workbook
from click.testing import CliRunner
from unittest.mock import patch, PropertyMock

from src.data.import_phenotypes import main as import_phenotypes
from src.features.smarterdb import Dataset, SampleSheep, Phenotype

from ..common import MongoMockMixin, SmarterIDMixin, SupportedChipMixin


class PhenotypeMixin(SmarterIDMixin, SupportedChipMixin, MongoMockMixin):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        # getting destination dataset
        cls.dst_dataset = Dataset.objects.get(file="test.zip")

        # create a src dataset
        cls.src_dataset = Dataset(
            file="test2.zip",
            country="Italy",
            species="Sheep",
            contents=[
                "phenotypes.xlsx"
            ]
        )
        cls.src_dataset.save()

        # create a workbook
        cls.workbook = Workbook()
        cls.sheet = cls.workbook.active

        # adding header
        cls.sheet.cell(row=1, column=1, value="Code")
        cls.sheet.cell(row=1, column=2, value="Name")
        cls.sheet.cell(row=1, column=3, value="Country")
        cls.sheet.cell(row=1, column=4, value="Id")
        cls.sheet.cell(row=1, column=5, value="ChestGirth")
        cls.sheet.cell(row=1, column=6, value="Height")
        cls.sheet.cell(row=1, column=7, value="Length")
        cls.sheet.cell(row=1, column=8, value="WidthOfPinBones")
        cls.sheet.cell(row=1, column=9, value="FAMACHA")
        cls.sheet.cell(row=1, column=10, value="Purpose")
        cls.sheet.cell(row=1, column=11, value="Alias")

        # adding values
        cls.sheet.cell(row=2, column=1, value="TEX")
        cls.sheet.cell(row=2, column=2, value="Texel")
        cls.sheet.cell(row=2, column=3, value="Italy")
        cls.sheet.cell(row=2, column=4, value="test-1")
        cls.sheet.cell(row=2, column=5, value=77.5)
        cls.sheet.cell(row=2, column=6, value=60.5)
        cls.sheet.cell(row=2, column=7, value=69.5)
        cls.sheet.cell(row=2, column=8, value=16)
        cls.sheet.cell(row=2, column=9, value="D")
        cls.sheet.cell(row=2, column=10, value="milk")
        cls.sheet.cell(row=2, column=11, value="alias-1")

    @classmethod
    def tearDownClass(cls):
        Dataset.objects.delete()
        SampleSheep.objects.delete()

        super().tearDownClass()

    def setUp(self):
        self.runner = CliRunner()

        # need also a sample
        self.sample = SampleSheep(
            original_id="test-1",
            smarter_id="ITOA-TEX-000000001",
            country="Italy",
            species="Sheep",
            breed="Texel",
            breed_code="TEX",
            dataset=self.dst_dataset,
            chip_name=self.chip_name,
            alias="alias-1",
        )
        self.sample.save()

    def tearDown(self):
        SampleSheep.objects.delete()

        super().tearDown()


class TestImportPhenotypeCLI(PhenotypeMixin, unittest.TestCase):
    def test_help(self):
        result = self.runner.invoke(import_phenotypes, ["--help"])
        self.assertEqual(0, result.exit_code)
        self.assertIn('Usage: main', result.output)


class TestImportPhenotypeByBreeds(PhenotypeMixin, unittest.TestCase):
    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_phenotype(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/phenotypes.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            result = self.runner.invoke(
                import_phenotypes,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "phenotypes.xlsx",
                    "--breed_column",
                    "Name",
                    "--purpose_column",
                    "Purpose",
                    "--chest_girth_column",
                    "ChestGirth",
                    "--height_column",
                    "Height",
                    "--length_column",
                    "Length",
                    "--additional_column",
                    "FAMACHA",
                    "--additional_column",
                    "WidthOfPinBones"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)
            self.sample.reload()
            self.assertIsInstance(self.sample.phenotype, Phenotype)

            reference = Phenotype(
                purpose="Milk",
                chest_girth=77.5,
                height=60.5,
                length=69.5,
                widthofpinbones=16,
                famacha="D"
            )

            self.assertEqual(reference, self.sample.phenotype)

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_phenotype_update(self, my_working_dir):
        """Test updating an existing genotype"""

        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/phenotypes.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            # assign a test phenotype with an attribute to be update and
            # another which won't be updated
            self.sample.phenotype = Phenotype(
                purpose="Milk",
                chest_girth=99.9999
            )
            self.sample.save()

            result = self.runner.invoke(
                import_phenotypes,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "phenotypes.xlsx",
                    "--breed_column",
                    "Name",
                    "--chest_girth_column",
                    "ChestGirth",
                    "--height_column",
                    "Height",
                    "--length_column",
                    "Length",
                    "--additional_column",
                    "FAMACHA",
                    "--additional_column",
                    "WidthOfPinBones"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)
            self.sample.reload()
            self.assertIsInstance(self.sample.phenotype, Phenotype)

            reference = Phenotype(
                purpose="Milk",
                chest_girth=77.5,
                height=60.5,
                length=69.5,
                widthofpinbones=16,
                famacha="D"
            )

            self.assertEqual(reference, self.sample.phenotype)

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_phenotype_multiple_sheet(self, my_working_dir):
        """Test updating an existing genotype"""

        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # add a new sheet to template
            new_sheet = self.workbook.create_sheet(title="colour")

            new_sheet.cell(row=1, column=1, value="Code")
            new_sheet.cell(row=1, column=2, value="Name")
            new_sheet.cell(row=1, column=3, value="Country")
            new_sheet.cell(row=1, column=4, value="Id")
            new_sheet.cell(row=1, column=5, value="Coat color")

            # adding values
            new_sheet.cell(row=2, column=1, value="TEX")
            new_sheet.cell(row=2, column=2, value="Texel")
            new_sheet.cell(row=2, column=3, value="Italy")
            new_sheet.cell(row=2, column=4, value="test-1")
            new_sheet.cell(row=2, column=5, value="white")

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/phenotypes.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            # assign a test phenotype with an attribute to be update and
            # another which won't be updated
            self.sample.phenotype = Phenotype(
                purpose="Milk",
                chest_girth=99.9999
            )
            self.sample.save()

            result = self.runner.invoke(
                import_phenotypes,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "phenotypes.xlsx",
                    "--sheet_name",
                    "colour",
                    "--breed_column",
                    "Name",
                    "--additional_column",
                    "Coat color",
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)
            self.sample.reload()
            self.assertIsInstance(self.sample.phenotype, Phenotype)

            reference = Phenotype(
                purpose="Milk",
                chest_girth=99.9999,
                coat_color="White"
            )

            self.assertEqual(reference, self.sample.phenotype)


class TestImportPhenotypeByBreedsNotExists(PhenotypeMixin, unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        # override value (set a breed which notx exist in database)
        cls.sheet.cell(row=2, column=1, value="MER")
        cls.sheet.cell(row=2, column=2, value="Merino")

    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_phenotype_raise_exc(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/phenotypes.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            result = self.runner.invoke(
                import_phenotypes,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "phenotypes.xlsx",
                    "--breed_column",
                    "Name",
                    "--purpose_column",
                    "Purpose",
                    "--chest_girth_column",
                    "ChestGirth",
                    "--height_column",
                    "Height",
                    "--length_column",
                    "Length",
                    "--additional_column",
                    "FAMACHA",
                    "--additional_column",
                    "WidthOfPinBones"
                ]
            )

            self.assertEqual(1, result.exit_code, msg=result.exception)
            self.sample.reload()
            self.assertIsNone(self.sample.phenotype)


class TestImportPhenotypeBySamples(PhenotypeMixin, unittest.TestCase):
    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_phenotype(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/phenotypes.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            result = self.runner.invoke(
                import_phenotypes,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "phenotypes.xlsx",
                    "--id_column",
                    "Id",
                    "--purpose_column",
                    "Purpose",
                    "--chest_girth_column",
                    "ChestGirth",
                    "--height_column",
                    "Height",
                    "--length_column",
                    "Length",
                    "--additional_column",
                    "FAMACHA",
                    "--additional_column",
                    "WidthOfPinBones"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)
            self.sample.reload()
            self.assertIsInstance(self.sample.phenotype, Phenotype)

            reference = Phenotype(
                purpose="Milk",
                chest_girth=77.5,
                height=60.5,
                length=69.5,
                widthofpinbones=16,
                famacha="D"
            )

            self.assertEqual(reference, self.sample.phenotype)


class TestImportPhenotypeByAlias(PhenotypeMixin, unittest.TestCase):
    @patch('src.features.smarterdb.Dataset.working_dir',
           new_callable=PropertyMock)
    def test_import_phenotype(self, my_working_dir):
        # create a temporary directory using the context manager
        with tempfile.TemporaryDirectory() as tmpdirname:
            working_dir = pathlib.Path(tmpdirname)
            my_working_dir.return_value = working_dir

            # save worksheet in temporary folder
            self.workbook.save(f"{working_dir}/phenotypes.xlsx")

            # got first sample from database
            self.assertEqual(SampleSheep.objects.count(), 1)

            result = self.runner.invoke(
                import_phenotypes,
                [
                    "--src_dataset",
                    "test2.zip",
                    "--dst_dataset",
                    "test.zip",
                    "--datafile",
                    "phenotypes.xlsx",
                    "--alias_column",
                    "Alias",
                    "--purpose_column",
                    "Purpose",
                    "--chest_girth_column",
                    "ChestGirth",
                    "--height_column",
                    "Height",
                    "--length_column",
                    "Length",
                    "--additional_column",
                    "FAMACHA",
                    "--additional_column",
                    "WidthOfPinBones"
                ]
            )

            self.assertEqual(0, result.exit_code, msg=result.exception)
            self.sample.reload()
            self.assertIsInstance(self.sample.phenotype, Phenotype)

            reference = Phenotype(
                purpose="Milk",
                chest_girth=77.5,
                height=60.5,
                length=69.5,
                widthofpinbones=16,
                famacha="D"
            )

            self.assertEqual(reference, self.sample.phenotype)
